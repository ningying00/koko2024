from openpyxl import load_workbook

from Settings.Config import testDataPath


class ExcelParse:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def load_workbook(self, filename):
        """
        加载数据文件
        :param filename:
        :return:
        """
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            raise e

    def get_sheet_by_name(self, sheetname):
        """
        获取sheet对象
        :param sheetname:
        :return:
        """
        try:
            self.sheet = self.workbook[sheetname]
        except Exception as e:
            raise e

    def get_cell_value(self, row, column):
        """
        获取单位格的值
        :param row:
        :param column:
        :return:
        """
        try:
            cell_value = self.sheet.cell(row=row, column=column).value
            return cell_value
        except Exception as e:
            raise e

    def get_row_value(self, row):
        """

        :param row:
        :return:
        """
        try:
            columns = self.sheet.max_column
            row_data = []
            for i in range(1, columns + 1):
                cell_value = self.sheet.cell(row=row, column=i).value
                row_data.append(cell_value)
            return row_data
        except Exception as e:
            raise e

    def get_rows_nums(self):
        """
        获取行数
        :return:
        """
        return self.sheet.max_row

    def get_columns_nums(self):
        """
        获取行数
        :return:
        """
        return self.sheet.max_column


if __name__ == '__main__':
    excel = ExcelParse()
    excel.load_workbook(testDataPath)
    excel.get_sheet_by_name("login")
    print(excel.get_columns_nums())
    print(excel.get_rows_nums())
    print(excel.get_row_value(1))
