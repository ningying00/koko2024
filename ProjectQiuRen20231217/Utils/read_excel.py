from openpyxl import load_workbook

from Settings.config import excelFilePath


class ReadExcel:
    def __init__(self):
        self.wb = None
        self.sheet = None

    def load_excel(self, filename):
        """
        打开excel文件
        :param filename: excel文件路径
        :return:
        """
        self.wb = load_workbook(filename)

    def sheet_names(self):
        """
        获取所有工作表名称
        :return:
        """
        sheet_names = self.wb.sheetnames
        return sheet_names

    def get_sheet(self, sheetname):
        self.sheet = self.wb[sheetname]
        #return self.sheet

    def get_row(self, i):
        """
        获取指定行，行对象
        :param i:
        :return:
        """
        row = self.sheet[i]
        return row

    def get_col(self, i):
        """
        获取制定列，列对象
        :param i:
        :return:
        """
        col = self.sheet[i]
        return col
    def get_row_max(self):
        return self.sheet.max_row

    def get_max_col(self):
        return self.sheet.max_column

    def get_row_value(self, i):
        """
        获取某一行的值
        :param i:
        :return:
        """
        columns = self.sheet.max_column
        row_list = []
        for j in range(1, columns + 1):
            value = self.sheet.cell(row=i, column=j).value
            row_list.append(value)
        return row_list

    def get_col_value(self, j):
        """
        获取某一列的值
        :param j:
        :return:
        """
        rows = self.sheet.max_row
        col_list = []
        for i in range(1, rows + 1):
            value = self.sheet.cell(row=i, column=j).value
            col_list.append(value)
        return col_list


if __name__ == '__main__':
    wb = ReadExcel()
    wb.load_excel(excelFilePath)
    wb.get_sheet("login")
    print(wb.get_row_value(1))
    print(wb.get_col_value(1))

